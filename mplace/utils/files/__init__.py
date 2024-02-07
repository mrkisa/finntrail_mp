import csv

from openpyxl import load_workbook


def read_excel(excel_file, sheet_index=0):
    wb = load_workbook(excel_file)
    sh = wb.worksheets[sheet_index]

    for value in sh.iter_rows(
            min_row=2, values_only=True):
        yield value


def read_csv(csv_file, delimiter=';'):
    with open(csv_file, 'r', encoding='utf8') as f:
        reader = csv.reader(f, delimiter=delimiter)
        next(reader)  # skip headers
        yield from reader
